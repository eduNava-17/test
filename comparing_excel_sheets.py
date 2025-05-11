import openpyxl
from openpyxl.styles import PatternFill

data_file1 = openpyxl.load_workbook(r'D:\datafile1.xlsx')
data_file2 = openpyxl.load_workbook(r'D:\datafile2.xlsx')

fill_style = PatternFill(start_color= 'FDD835',end_color='FDD835',fill_type='lightHorizontal')

data_sheet1 = data_file1['Sheet1']
data_sheet2 = data_file2['Sheet1']

for row in data_sheet1.iter_rows():
    for cell in row:
        cell_value = cell.value

        if cell_value != data_sheet2[cell.coordinate].value:
            cell.fill = fill_style

data_file1.save(r'D:\compared_file.xlsx')
